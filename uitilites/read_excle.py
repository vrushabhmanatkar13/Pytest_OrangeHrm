import os.path
from typing import final, Final

import openpyxl
import pytest


# install openpyxl package


def load_excel_file(excel_path):
    global workbook
    path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    workbook = openpyxl.load_workbook(path + excel_path)


def get_data(sheet_name, testcase_name):
    load_excel_file("/TestData/Data.xlsx")
    sheet = workbook[sheet_name]
    row = sheet.max_row
    col = sheet.max_column
    current_testcase = ""
    data_list = []
    for r in range(1, row + 1):
        testcase = sheet.cell(r, 1).value
        if testcase == testcase_name:
            current_testcase = sheet.cell(r, 1).value  # Get the name of the new test case
        elif current_testcase and testcase == "***":
            break
        elif current_testcase and testcase is not None:
            row_list = []
            for c in range(1, col + 1):
                value = sheet.cell(r + 1, c).value
                if value is None or value == "***":
                    break
                row_list.append(value)
            if row_list:
                data_list.append(row_list)
    return data_list


def get_data_by_index(sheet_name, row_index):
    load_excel_file("/TestData/Data.xlsx")
    sheet = workbook[sheet_name]
    column = sheet.max_column
    final_data = []
    for c in range(1, column + 1):
        data = sheet.cell(row_index, c).value
        if data is None:
            break
        final_data.append(data)
    return final_data


def get_data_by_key(sheet_name, testcase_name, row_num):
    load_excel_file("/TestData/Data.xlsx")
    sheet = workbook[sheet_name]
    row = sheet.max_row
    column = sheet.max_column
    current_test = ""
    final_data = {}
    for r in range(1, row + 1):
        testcase = sheet.cell(r, 1).value
        if testcase == testcase_name:
            current_test = sheet.cell(r, 1).value  # Get the name of the new test case
            a: Final = r
        elif current_test and testcase == "***":
            break
        elif current_test and testcase is not None:
            for c in range(1, column + 1):
                key = sheet.cell(a + 1, c).value
                data = sheet.cell(a + 1 + row_num, c).value
                if data is None or data == "***":
                    break
                final_data[key] = data
    return final_data

